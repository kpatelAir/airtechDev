import os
import pandas as pd
from urllib.parse import urlparse
import openpyxl
import requests

def replace_empty_cells(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    max_row = sheet.max_row

    # Iterate through each row
    for row in range(2, max_row + 1):
        # Check if column B is not empty
        if sheet.cell(row=row, column=2).value:
            # Replace empty cells in columns C and D with 'N/A'
            if sheet.cell(row=row, column=3).value is None:
                sheet.cell(row=row, column=3).value = 'N/A'
            if sheet.cell(row=row, column=4).value is None:
                sheet.cell(row=row, column=4).value = 'N/A'

    # Save the modified workbook
    wb.save(file_path)
    print(f"Empty cells in columns C and D replaced with 'N/A' where column B is not empty in {file_path}")

def read_text_file(input_file):
    data = []
    with open(input_file, 'r') as file:
        for line in file:
            if "ID:" in line and "Company:" in line and "Website:" in line:
                id_str, rest = line.split(', ', 1)
                company_str, website_str = rest.split(', Website: ')
                
                # Extracting ID, Company Name, and Website
                company_id = id_str.split(': ')[1]
                company_name = company_str.split(': ')[1]
                company_website = website_str.rstrip()

                data.append({'ID': company_id, 'Company Name': company_name, 'Website': company_website})
    return data

def write_to_excel(output_file, data):
    df = pd.DataFrame(data)
    df.to_excel(output_file, index=False)
    print(f"Data written to {output_file}")


def extract_domain(url):
    # Parse the URL
    parsed_url = urlparse(url)

    # Get the netloc part (domain)
    domain = parsed_url.netloc

    # Remove 'www.' if present
    if domain.startswith('www.'):
        domain = domain[4:]

    return domain

def process_excel(input_file, output_file):
    # Read Excel file into a DataFrame
    df = pd.read_excel(input_file)

    # Convert 'Website' column values to strings before applying extract_domain
    df['Domain'] = df['Website'].astype(str).apply(extract_domain)

    # Save the DataFrame with the new column to a new Excel file
    df.to_excel(output_file, index=False)


def add_www_to_domains(excel_file):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    
    # Iterate through the rows starting from the second row
    for row in range(2, sheet.max_row + 1):
        print(f"Starting row {row}")
        domain = sheet[f'D{row}'].value
        if domain == 'N/A':
            sheet[f'E{row}'] = 'N/A'
            continue  # Move to the next row
        
        if domain:
            domain_with_www = f'www.{domain}'
            try:
                # Test if the domain with www. is accessible with a timeout of 10 seconds
                response = requests.get(f'http://{domain_with_www}', timeout=10)
                if response.status_code == 200:
                    sheet[f'E{row}'] = domain_with_www
                else:
                    sheet[f'E{row}'] = domain
            except requests.RequestException:
                sheet[f'E{row}'] = domain
        
        # Save the changes to the Excel file every 25 rows
        if row % 25 == 0:
            wb.save(excel_file)
            print("Saved progress.")
    
    # Save the changes to the Excel file after processing all rows
    wb.save(excel_file)
    print("Process completed.")






if __name__ == "__main__":
   
    '''
    # Text file to Excel
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    input_file = os.path.join(desktop_path, 'company_data.txt')
    output_file = os.path.join(desktop_path, 'company_data.xlsx')
    extracted_data = read_text_file(input_file)
    write_to_excel(output_file, extracted_data)

    # Create Domain column
    input_file = os.path.join(desktop_path, 'company_data.xlsx')
    output_file = os.path.join(desktop_path, 'company_data.xlsx')
    process_excel(input_file, output_file)
    print(f"Domain extraction complete. Output saved to {output_file}")
  
    # Insert N/A in empty cells
    replace_empty_cells('company_data.xlsx')
    '''

    # Add www.Domain column
    add_www_to_domains('company_data.xlsx')



