import whois
import dns.resolver
from googlesearch import search
from fuzzywuzzy import fuzz
from urllib.parse import urlparse
import os
import openpyxl
from time import sleep

# --- Enhanced Scoring Based on Multiple Heuristics ---
def score_url(company_name, url):
    header_text = None
    #header_text = scrape_header(url)
    #print(header_text)
    
    domain = urlparse(url).netloc.lower().replace("www.", "")
    score = 0

    # Preferred TLDs
    preferred_tlds = [".com", ".org", ".edu", ".gov", ".net"]
    if any(domain.endswith(tld) for tld in preferred_tlds):
        score += 20

    # Penalize suspicious subdomains (e.g., 'blog.', 'shop.')
    if domain.startswith(("blog.", "shop.", "news.")):
        score -= 10

    if header_text:
        header_similarity_score = fuzz.partial_ratio(company_name.lower(), header_text.lower())
        print(f"Header similarity score for {company_name.lower()} in {url}: {header_similarity_score}")
        score += header_similarity_score
    return score

def google_search_for_website(company_name, postal_code = None):
    query = f"{company_name} {postal_code} official website"
    try:
        candidates = list(search(query, num_results=3))
        if not candidates:
            return None
        # Score each URL and pick the best
        ranked_candidates = sorted(
            candidates, key=lambda x: score_url(company_name, x), reverse=True
        )
        return ranked_candidates[0]  # Best match based on score
    except Exception as e:
        print(f"Error: {e}")
        return None

def get_official_website(company_name, postalCode):
    # Search Google for official website
    website = google_search_for_website(company_name, postalCode)
    if website:
        print(f"{company_name}: {website}")
        return website


columnMap = {
    'CompanyName': 2,
    'SaveSite' : 3,
    'PostalCode': 4,
    'City' : 5,
}

if __name__ == "__main__":
    # Change file name
    file_name = "hubspot-crm-exports-company_site-2025-02-26.xlsx"
    
    # Assuming the file is on the desktop
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    # Full path to the Excel file
    file_path = os.path.join(desktop_path, file_name)

    workbook = openpyxl.load_workbook(file_path)
    print("Excel file opened successfully!")
    # Access the active sheet
    sheet = workbook.active
    max_row = 10 # use sheet.max_row for the full sheet

    for row in range(2, max_row + 1):
        # Save the first column cell value
        companyName = sheet.cell(row=row, column=columnMap['CompanyName']).value
        postalCode = sheet.cell(row=row, column=columnMap['PostalCode']).value

        site = get_official_website(companyName, postalCode)

        sheet.cell(row=row, column=columnMap['SaveSite']).value = site

        if row % 25 == 0:
            sheet.save(file_path)
            print("Saved progress.")

        sleep.delay(3) # Delay for 3 seconds to avoid being blocked by Google
    
    print("Rows complete.")
    # Save the changes to the workbook
    workbook.save(file_path)
    print("Excel file updated successfully!")


