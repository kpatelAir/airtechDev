import os
import re
import requests
import openpyxl
from bs4 import BeautifulSoup
from collections import Counter
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
import cloudscraper
import socket
import random

""" README Section: 

This script is used to scrape the text content of a given URL and determine the top 3 ECC categories based on the keywords provided. 
If the top category is not found, the script will attempt to scrape the page using Selenium and return the top category. 

"""




# Define the codes, categories, and keywords
categories = {
    'ECC1': 'Aerospace, Aero, Air, Aircraft, Aviation, Boeing, Plane, Planes, Airplane, Airplanes, Jet, Aerodynamics, Hypersonics, Helicopter, Flight, Space, Spacecraft, Boosters, Rocket, Rockets, Skymiles, Check-in, Trips, Travel, Radomes, Antennas, Satellite',
    'ECC2': 'Marine, Naval, Maritime, Ocean, Yacht, Aquatic, Vessel, Boat, Boats, Lake, River, Powerboat, RIB, RIBs, Catamaran, Sailing',
    'ECC3': 'PCB, Circuit, Multilayer, Soldering, Electronic',
    'ECC4': 'Wind, Turbine, Green, Clean',
    'ECC5': 'Automotive, Car, Cars, Vehicle, Vehicles, Tire, Tires, Truck, Trucks, Race, Racing, Nascar, Motorcycles, Motorcycle, Handling, Ride',
    'ECC6': 'UAV, Drone, Aerial, Surveillance, Robotics, Mapping, Autonomous, Remote',
    'ECC7': 'Glass, Transparent, Glazing, Decorative',
    'ECC8': 'Architecture, Architectural, Urban, Landscape, Construction',
    'ECC9': 'General, Composites, Composite, Polymer, Polymers, Fiber, Fibers, Bagging, Tape, Tubing, Shrink, Additive, Fabrication',
    'ECC10': 'Intercompany, Acquisition, Merger, Partnership, Partners, Alliance, Integration, Entities, Divisions, Businesses, Companies, Investor Relations, Our Brands',
    'ECC11': '3D, Print, Tooling, Prototype, Mold',
    'ECC12': '3D, Print, Resin, Additive, Printing, Curing, Printers',
    'ECC13': 'Army, Air Force, Marines, Troops', # Military
    'ECC14': 'Education, Academic, Scholars, Graduate, Alumni',
    'ECC15': 'Supplier, Supply, Chain, Supplies, Procurement, Clearance, Products, Distribution, Promotions',
    'ECC16': 'Medical, Healthcare, Medicine, Drug, Molecule, Lab, Health, Healthier',
}

def get_free_port():
    # Create a temporary socket to find a free port
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.bind(('', 0))
    addr, port = s.getsockname()
    s.close()
    return port

port = get_free_port()
print(f"Using ChromeDriver on random port: {port}")

# Initialize the Chrome driver
service = ChromeService(executable_path=ChromeDriverManager().install(), port=port)
options = webdriver.ChromeOptions()
options.add_argument("--headless=new")
options.add_argument("--disable-extensions")
options.add_argument("--disable-gpu")

driver = webdriver.Chrome(service=service, options=options)

# Function to clean and process text
def clean_text(text):
    # Remove special characters and numbers, keep only letters and spaces
    clean_text = re.sub(r'[^a-zA-Z\s]', '', text)
    return clean_text.lower()


# Scrape given URL for all text available and return single string all text
def scrape(url): 
    # Define the User-Agent to avoid error 403
    headers = requests.utils.default_headers()
    headers.update(
        {
            'User-Agent': 'ECC Agent 1.0',
            'From': 'kpatel@airtech.com'
        }
    )
    try:
        # Send a GET request to the URL
        response = requests.get(url, headers=headers, timeout=(10, 10))

        # Check if the request was successful
        if response.status_code == 200:
            # Parse the HTML content of the page
            soup = BeautifulSoup(response.text, 'lxml')
            
            # Find all text content
            texts = soup.stripped_strings
            
            # Combine all text into a single string
            all_text = ' '.join(texts)
            all_text = clean_text(all_text)

            #print(all_text) # Testing print
            return all_text
        
        else:
            print(f"Failed to retrieve the page. Status code: {response.status_code}")
            scraper = cloudscraper.CloudScraper()  # CloudScraper for one more try at scrape
            if scraper:
                return scraper.get(url).text
            else:
                return ' '

    except requests.exceptions.Timeout:
        print(f"Request to {url} timed out.")
        return ' '
    except requests.exceptions.RequestException as e:
        print(f"Request to {url} failed: {e}")
        return ' '


def scrape_header(url):
    # Define the User-Agent to avoid error 403
    headers = requests.utils.default_headers()
    headers.update( {
            'User-Agent': 'ECC Agent 1.0',
            'From': 'kpatel@airtech.com'
        }
    )
    try:
        # Send a GET request to the URL
        response = requests.get(url, headers=headers, timeout=(10, 10))

        # Check if the request was successful
        if response.status_code == 200:
            # Parse the HTML content of the page
            soup = BeautifulSoup(response.text, 'lxml')
            
            # Find the header section
            header = soup.find('header')
            
            if header:
                # Extract the text content from the header
                header_text = ' '.join(header.stripped_strings)
                return header_text
            else:
                print("Header not found.")
                return None
        else:
            print(f"Failed to retrieve the page. Status code: {response.status_code}")
            return None

    except requests.exceptions.Timeout:
        print(f"Request to {url} timed out.")
        return None
    except requests.exceptions.RequestException as e:
        print(f"Request to {url} failed: {e}")
        return None

# Function to get ECC category counts from string and scrape JS sites
def get_ecc(text, url=None):
    if text != ' ':
        # Ensure text is clean
        clean_text_content = clean_text(text)
        
        # Create a counter for the categories
        category_counts = Counter()

        # Iterate through each category and its keywords
        for ecc_code, keywords in categories.items():
            # Split keywords by comma and strip any extra whitespace
            keyword_list = [keyword.strip().lower() for keyword in keywords.split(',')]
            
            # Count occurrences of each keyword in the cleaned text
            for keyword in keyword_list:
                # Use regular expression to find whole words
                pattern = rf'\b{keyword}\b'
                count = len(re.findall(pattern, clean_text_content))
                category_counts[ecc_code] += count

        # Get the top 3 categories by count
        top_categories = category_counts.most_common(3)

       # If no keywords found and URL is provided, scrape the page using Selenium this time
        if top_categories[0][1] == 0 and url:
            try:
                if url:
                    driver.get(url)
                    soup = BeautifulSoup(driver.page_source, 'lxml')
                    texts = soup.stripped_strings
                    all_text = ' '.join(texts)
                    all_text = clean_text(all_text)
                    return get_ecc(all_text)  # Recursive call without the URL to avoid infinite recursion
                else:
                    return 'None'  # Return None if there's no URL provided
            except requests.exceptions.Timeout:
                print(f"Request 2 to {url} timed out.")
                return 'None'  # Return None if the request times out
            except Exception as e:
                print(f"An error occurred while processing {url}: {e}")
                return 'None'
        else:
            # Return the top categories
            return top_categories if top_categories else None
    else:
        return 'None'
  

def ECC (file_path, url_column, write_column):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        print("Excel file opened successfully!")
        # Access the active sheet
        sheet = workbook.active
        max_row = sheet.max_row

        # Iterate through each row starting from the second row
        for row in range(2, max_row + 1):
            # Save the first column cell value
            url = sheet.cell(row=row, column=url_column).value
            if url:
                if "aero" in url:
                    sheet.cell(row=row, column=write_column).value = "ECC1"
                    print(f"{row}.{url}: ECC1")
                elif "yacht" in url or "boat" in url:
                    sheet.cell(row=row, column=write_column).value = "ECC2"
                    print(f"{row}.{url}: ECC2")
                elif "racing" in url or "automotive" in url:
                    sheet.cell(row=row, column=write_column).value = "ECC5"
                    print(f"{row}.{url}: ECC5")
                elif "glass" in url:
                    sheet.cell(row=row, column=write_column).value = "ECC7"
                    print(f"{row}.{url}: ECC7")
                elif "composite" in url:
                    sheet.cell(row=row, column=write_column).value = "ECC9"
                    print(f"{row}.{url}: ECC9")
                elif ".mil/" in url:
                    sheet.cell(row=row, column=write_column).value = "ECC13"
                    print(f"{row}.{url}: ECC13")
                elif ".edu/" in url:
                    sheet.cell(row=row, column=write_column).value = "ECC14"
                    print(f"{row}.{url}: ECC14")
                elif "sales" in url:
                    sheet.cell(row=row, column=write_column).value = "ECC15"
                    print(f"{row}.{url}: ECC15")
                elif "N/A" in url:
                    sheet.cell(row=row, column=write_column).value = "N/A"
                    print(f"{row}.{url}: N/A")

                else:
                    try:
                        # Scrape the website and get the ECC category
                        text = scrape(url)
                        top_ecc = get_ecc(text,url)
                        print(f"{row}.{url}: {top_ecc}")

                        # Write the top ECC category to the specified column
                        if top_ecc:
                            if top_ecc == 'None':
                                sheet.cell(row=row, column=write_column).value = "NONE"
                            elif top_ecc[0][1] <= 2:
                                scraper = cloudscraper.CloudScraper()  # CloudScraper for one more try at scrape
                                top_ecc = get_ecc((scraper.get(url).text))
                                print(f"{row}.{url}: {top_ecc}")
                                if top_ecc:
                                    if top_ecc == 'None':
                                        sheet.cell(row=row, column=write_column).value = "NONE"
                                    elif top_ecc[0][1] <= 1:
                                        sheet.cell(row=row, column=write_column).value = "Z" 
                                    else:
                                        sheet.cell(row=row, column=write_column).value = top_ecc[0][0]
                                else:
                                    sheet.cell(row=row, column=write_column).value = "ERROR"
                            else:
                                sheet.cell(row=row, column=write_column).value = top_ecc[0][0]
                        else:
                            sheet.cell(row=row, column=write_column).value = "ERROR"

                    except Exception as e:
                        print(f"An error occurred while processing {url}: {e}")
                        sheet.cell(row=row, column=write_column).value = "ERROR"
            else:
                sheet.cell(row=row, column=write_column).value = "EMPTY URL"

        print("All rows complete.")
        # Save the changes to the workbook
        workbook.save(file_path)
        print("Excel file updated successfully!")

    except Exception as e:
        print(f"An error occurred: {e}")


# Main
if __name__ == "__main__":
    """
    # Change file name
    file_name = "JasonsNAVSales.xlsx"
    
    # Assuming the file is on the desktop
    desktop_path = os.path.join(os.path.expanduser('~'), 'Desktop')
    # Full path to the Excel file
    file_path = os.path.join(desktop_path, file_name)

    # Run script
    # File path, URL column num, print column num
    ECC(file_path, 1, 26)
    """

    # Test the ECC function with a sample URL
    scrape_text = scrape("https://ame.usc.edu/")
    print(scrape_text)
    print(get_ecc(scrape_text, "https://ame.usc.edu/"))

    driver.close()
